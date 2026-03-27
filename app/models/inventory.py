import pandas as pd
from flask import current_app
from openpyxl import load_workbook
from datetime import datetime
import json
import os
import pickle


# ── 快取 ────────────────────────────────────────────────────────────
_INVENTORY_CACHE = {'mtime': 0, 'data': None}
_MASTER_MODEL_CACHE = {'mtime': 0, 'data': None}

def normalize_model_name(model_name):
    """標準化機型名稱，移除後綴變體"""
    import re
    if not model_name:
        return model_name
    
    normalized = str(model_name).strip()
    
    # 移除常見後綴（主底座、副底座、富底座）
    normalized = re.sub(r'(主底座|副底座|富底座)$', '', normalized)
    
    # 移除括號內容（如 (工作台-鑽孔)）
    normalized = re.sub(r'\([^)]*\)$', '', normalized)
    
    # 清理尾部的連字符、底線和空格
    normalized = re.sub(r'[-_\s]+$', '', normalized).strip()
    
    return normalized

def _get_master_model_list(xl=None):
    """從所有零件工作表收集完整機型清單（保留原始名稱，智能分組排序）"""
    global _MASTER_MODEL_CACHE
    try:
        casting_file = current_app.config['CASTING_FILE']
        mtime = os.path.getmtime(casting_file)

        # 若快取仍有效，直接回傳
        if _MASTER_MODEL_CACHE['mtime'] == mtime and _MASTER_MODEL_CACHE['data'] is not None:
            return _MASTER_MODEL_CACHE['data']

        # 嘗試讀取持久化快取
        cache_dir = os.path.join(os.getcwd(), 'app', 'cache')
        os.makedirs(cache_dir, exist_ok=True)
        cache_file = os.path.join(cache_dir, 'master_models.pkl')
        if os.path.exists(cache_file):
            try:
                with open(cache_file, 'rb') as f:
                    cached = pickle.load(f)
                if cached.get('mtime') == mtime:
                    _MASTER_MODEL_CACHE['mtime'] = mtime
                    _MASTER_MODEL_CACHE['data'] = cached['data']
                    return cached['data']
            except:
                pass

        if xl is None:
            # 使用更快的 calamine 引擎
            xl = pd.ExcelFile(casting_file, engine='calamine')

        all_models = {}

        # 從四個零件分頁收集所有機型，保留 Excel 原有的顯示順序
        for sheet_name, sheet_idx in [('底座', 1), ('工作台', 2), ('橫樑', 3), ('立柱', 4)]:
            try:
                df = pd.read_excel(xl, sheet_name=sheet_idx)
                # 機型通常在第二欄（index 1）
                if len(df.columns) > 1:
                    model_col = df.columns[1]
                    # unique() 會依序回傳遇到的元素，因此能保留 Excel 當中出現的上下順序
                    models = df[model_col].dropna().unique()
                    for m in models:
                        model_str = str(m).strip()
                        if model_str and model_str != 'nan' and model_str != '品號':
                            all_models[model_str] = True
            except Exception as e:
                print(f"Error reading {sheet_name}: {e}")

        # 取代智能排序，這裡直接讓它乖乖照著 Excel 原有的順序
        result = list(all_models.keys())

        # 更新快取
        _MASTER_MODEL_CACHE['mtime'] = mtime
        _MASTER_MODEL_CACHE['data'] = result
        
        try:
            with open(cache_file, 'wb') as f:
                pickle.dump({'mtime': mtime, 'data': result}, f)
        except:
            pass
            
        return result

    except Exception as e:
        print(f"Error loading master model list: {e}")
        return []

def load_casting_inventory():
    """載入鑄件庫存資料（含快取：只有 Excel 修改時才重新讀取）"""
    global _INVENTORY_CACHE
    try:
        casting_file = current_app.config['CASTING_FILE']
        mtime = os.path.getmtime(casting_file)

        # 若快取仍有效，直接回傳
        if _INVENTORY_CACHE['mtime'] == mtime and _INVENTORY_CACHE['data'] is not None:
            return _INVENTORY_CACHE['data']

        # 嘗試讀取持久化快取
        cache_dir = os.path.join(os.getcwd(), 'app', 'cache')
        cache_file = os.path.join(cache_dir, 'inventory_cache.pkl')
        if os.path.exists(cache_file):
            try:
                with open(cache_file, 'rb') as f:
                    cached = pickle.load(f)
                if cached.get('mtime') == mtime:
                    _INVENTORY_CACHE['mtime'] = mtime
                    _INVENTORY_CACHE['data'] = cached['data']
                    return cached['data']
            except:
                pass

        # 使用更快的 calamine 引擎 
        xl = pd.ExcelFile(casting_file, engine='calamine')

        
        # 從各個鑄件工作表計算總數
        inventory = {}
        semi_finished = {}  # 半品總數
        finished = {}  # 成品總數
        all_models = {}  # 儲存所有機型的數據

        # 定義半品和成品的欄位映射
        semi_finished_fields = {
            '底座': ['素材', 'M4', 'M3'],
            '工作台': ['素材', 'W1', 'W2', 'W3', 'W4'],
            '橫樑': ['素材', 'M6', 'M5'],
            '立柱': ['素材', '半品', '成品銑工']
        }
        
        finished_fields = {
            '底座': ['成品研磨'],
            '工作台': ['成品'],
            '橫樑': ['成品研磨'],
            '立柱': ['成品研磨']
        }

        # 1. 建立完整機型清單（從所有零件分頁收集）
        master_models = _get_master_model_list(xl=xl)
        for model_str in master_models:
            all_models[model_str] = {'機型': model_str, '工作台': 0, '底座': 0, '橫樑': 0, '立柱': 0}
        
        for part_name, sheet_idx in [('底座', 1), ('工作台', 2), ('橫樑', 3), ('立柱', 4)]:
            try:
                df = pd.read_excel(xl, sheet_name=sheet_idx)
                
                # 取得所有製程欄位的索引（排除總數欄位）
                config = CONFIGS.get(part_name, [])
                process_col_indices = [idx for label, idx in config if label != '總數']
                
                # 建立欄位名稱到索引的映射
                field_to_idx = {label: idx for label, idx in config}
                
                # 初始化該料件的總數
                part_total = 0
                part_semi = 0
                part_finished = 0

                # 提取機型詳細資料並累加總數
                if '機型' in df.columns:
                    for _, row in df.iterrows():
                        model = row.get('機型', '')
                        # 排除無效機型
                        if pd.notna(model) and str(model).strip() and str(model).strip() != '品號':
                            model_str = str(model).strip()
                            
                            # 初始化該機型的數據
                            if model_str not in all_models:
                                all_models[model_str] = {'機型': model_str, '工作台': 0, '底座': 0, '橫樑': 0, '立柱': 0}
                            
                            # 加總所有製程欄位的數量
                            row_total = 0
                            row_semi = 0
                            row_finished = 0
                            
                            for col_idx in process_col_indices:
                                if col_idx < len(row):
                                    val = row.iloc[col_idx]
                                    if pd.notna(val):
                                        try:
                                            qty = int(float(val))
                                            row_total += qty
                                        except:
                                            pass
                            
                            # 計算半品數量
                            for field_name in semi_finished_fields.get(part_name, []):
                                if field_name in field_to_idx:
                                    col_idx = field_to_idx[field_name]
                                    if col_idx < len(row):
                                        val = row.iloc[col_idx]
                                        if pd.notna(val):
                                            try:
                                                row_semi += int(float(val))
                                            except:
                                                pass
                            
                            # 計算成品數量
                            for field_name in finished_fields.get(part_name, []):
                                if field_name in field_to_idx:
                                    col_idx = field_to_idx[field_name]
                                    if col_idx < len(row):
                                        val = row.iloc[col_idx]
                                        if pd.notna(val):
                                            try:
                                                row_finished += int(float(val))
                                            except:
                                                pass
                            
                            # 累加到總表 (使用 += 處理重複出現的機型)
                            all_models[model_str][part_name] += row_total
                            
                            # 累加到該料件總數
                            part_total += row_total
                            part_semi += row_semi
                            part_finished += row_finished
                
                # 設定該料件的總庫存
                inventory[part_name] = part_total
                semi_finished[part_name] = part_semi
                finished[part_name] = part_finished
                
            except Exception as e:
                print(f"Error loading {part_name}: {e}")
                inventory[part_name] = 0
                semi_finished[part_name] = 0
                finished[part_name] = 0
        
        # 確保順序：工作台、底座、橫樑、立柱
        inventory = {
            '工作台': inventory.get('工作台', 0),
            '底座': inventory.get('底座', 0),
            '橫樑': inventory.get('橫樑', 0),
            '立柱': inventory.get('立柱', 0)
        }
        
        semi_finished = {
            '工作台': semi_finished.get('工作台', 0),
            '底座': semi_finished.get('底座', 0),
            '橫樑': semi_finished.get('橫樑', 0),
            '立柱': semi_finished.get('立柱', 0)
        }
        
        finished = {
            '工作台': finished.get('工作台', 0),
            '底座': finished.get('底座', 0),
            '橫樑': finished.get('橫樑', 0),
            '立柱': finished.get('立柱', 0)
        }
        
        # 轉換機型數據為列表 (包含零庫存機型)
        details = list(all_models.values())
        
        result = {
            'summary': inventory,
            'semi_finished': semi_finished,
            'finished': finished,
            'details': details,
            'all_models': all_models
        }

        # 更新快取
        _INVENTORY_CACHE['mtime'] = mtime
        _INVENTORY_CACHE['data'] = result
        
        try:
            with open(cache_file, 'wb') as f:
                pickle.dump({'mtime': mtime, 'data': result}, f)
        except:
            pass
            
        
        return result

    except Exception as e:
        print(f"Error loading casting inventory: {e}")
        return {'summary': {}, 'semi_finished': {}, 'finished': {}, 'details': [], 'all_models': {}, 'error': str(e)}

def update_history_note(part, item_id, timestamp, field, new_note):
    """更新歷史紀錄中的備註資訊"""
    try:
        log_file = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'logs', 'edit_history.json')
        
        if not os.path.exists(log_file):
            return False
        
        with open(log_file, 'r', encoding='utf-8') as f:
            history = json.load(f)
        
        updated = False
        for h in history:
            # 使用 part, item_id, timestamp, field 作為複合鍵尋找紀錄
            if (h.get('part') == part and 
                str(h.get('item_id')) == str(item_id) and 
                h.get('timestamp') == timestamp and 
                h.get('field') == field):
                
                h['note'] = new_note
                updated = True
                break
        
        if updated:
            with open(log_file, 'w', encoding='utf-8') as f:
                json.dump(history, f, ensure_ascii=False, indent=2)
            return True
            
        return False
    
    except Exception as e:
        print(f"Error updating history note: {e}")
        return False

def get_zero_inventory_models():
    """獲取總數為0的機型列表（庫存警示）"""
    try:
        casting_data = load_casting_inventory()
        all_models = casting_data.get('all_models', {})
        
        # 找出總數為0的機型
        zero_models = []
        for model_str, model_data in all_models.items():
            total = sum(model_data.get(part, 0) for part in ['工作台', '底座', '橫樑', '立柱'])
            if total == 0:
                zero_models.append({
                    '機型': model_str,
                    '工作台': model_data.get('工作台', 0),
                    '底座': model_data.get('底座', 0),
                    '橫樑': model_data.get('橫樑', 0),
                    '立柱': model_data.get('立柱', 0)
                })
        
        return zero_models
    except Exception as e:
        print(f"Error getting zero inventory models: {e}")
        return []


# 欄位配置 (全域，供多個函式使用)
CONFIGS = {
    '底座': [('素材', 2), ('M4', 3), ('M3', 4), ('成品研磨', 5), ('總數', 7)],
    '工作台': [('素材', 2), ('W1', 3), ('W2', 4), ('W3', 5), ('W4', 6), ('成品', 7), ('總數', 9)],
    '橫樑': [('素材', 2), ('M6', 4), ('M5', 5), ('成品研磨', 6), ('總數', 8)],
    '立柱': [('素材', 2), ('半品', 3), ('成品銑工', 4), ('成品研磨', 5), ('總數', 6)]
}

SHEET_MAP = {'底座': 1, '工作台': 2, '橫樑': 3, '立柱': 4}

def get_part_details(part_type):
    """載入特定鑄件的詳細製程資料 (顯示所有原始機型名稱)"""
    try:
        casting_file = current_app.config['CASTING_FILE']
        sheet_idx = SHEET_MAP.get(part_type)
        if sheet_idx is None: return {"headers": [], "rows": []}

        config = CONFIGS.get(part_type, [])
        headers = ['品號', '機型'] + [c[0] for c in config]
        
        df = pd.read_excel(casting_file, sheet_name=sheet_idx)
        
        # 建立現有數據索引
        existing_rows = {}
        for _, row in df.iterrows():
            if pd.isna(row.iloc[0]) and pd.isna(row.iloc[1]): 
                continue
            model_name = str(row.iloc[1]).strip()
            existing_rows[model_name] = row

        def to_int(val):
            try: 
                return int(float(val)) if pd.notna(val) else 0
            except: 
                return 0
                
        rows = []
        for model_name, row in existing_rows.items():
            part_number = str(row.iloc[0]).strip()
            
            # 處理浮點數被轉成字串的 .0 尾巴 (例如 "12345.0" -> "12345")
            if part_number.endswith('.0'):
                part_number = part_number[:-2]
            
            # 若品號為空、nan 或 N/A，改回空字串，但不跳過該機型
            if part_number in ['nan', 'N/A', 'None']:
                part_number = ''
                
            data_row = {'機型': model_name, '品號': part_number}
            for label, idx in config:
                data_row[label] = to_int(row.iloc[idx])
            rows.append(data_row)
            
        return {"headers": headers, "rows": rows}
    except Exception as e:
        print(f"Error loading {part_type} details: {e}")
        import traceback
        traceback.print_exc()
        return {"headers": [], "rows": []}


def update_cell(part_type, item_id, field, new_value, user_id, model_name=None):
    """更新 Excel 儲存格並記錄歷程 (支援依機型更新及自動新增列)"""
    try:
        casting_file = current_app.config['CASTING_FILE']
        sheet_idx = SHEET_MAP.get(part_type)
        config = CONFIGS.get(part_type, [])
        
        if sheet_idx is None:
            return {'success': False, 'error': '無效的鑄件類型'}
        
        # 找出欄位索引
        col_idx = None
        for label, idx in config:
            if label == field:
                col_idx = idx
                break
        
        if col_idx is None:
            return {'success': False, 'error': f'無效的欄位: {field}'}
        
        # 使用 openpyxl 開啟並修改
        wb = load_workbook(casting_file)
        ws = wb.worksheets[sheet_idx]
        
        # 找出對應行
        target_row = None
        old_value = None
        
        # 方式1: 依品號搜尋 (僅當 item_id 不是 'N/A' 時)
        if item_id and item_id != 'N/A':
            for row_num in range(2, ws.max_row + 1):
                cell_value = str(ws.cell(row=row_num, column=1).value).strip()
                if cell_value.endswith('.0'): cell_value = cell_value[:-2]
                
                check_id = str(item_id).strip()
                if check_id.endswith('.0'): check_id = check_id[:-2]
                
                if cell_value == check_id:
                    target_row = row_num
                    break
        
        
        # 方式2: 依機型搜尋 (如果 item_id 找不到或為 'N/A')
        if target_row is None and model_name:
            for row_num in range(2, ws.max_row + 1):
                model_cell_value = ws.cell(row=row_num, column=2).value
                if model_cell_value and str(model_cell_value).strip() == str(model_name).strip():
                    target_row = row_num
                    # 同時更新 item_id 以便後續記錄
                    item_id = ws.cell(row=row_num, column=1).value or 'N/A'
                    break
        
        # 如果還是找不到，且有機型名稱，則新增一行
        if target_row is None and model_name:
            target_row = ws.max_row + 1
            ws.cell(row=target_row, column=1, value='N/A')  # 新增列品號設為 N/A
            ws.cell(row=target_row, column=2, value=model_name)
            item_id = 'N/A'
            print(f"[INFO] Appended new row for model: {model_name}")
        
        if target_row is None:
            return {'success': False, 'error': '找不到對應的品號或機型'}
        
        # 取得舊值
        old_value = ws.cell(row=target_row, column=col_idx + 1).value
        
        # 更新儲存格 (openpyxl column is 1-indexed)
        ws.cell(row=target_row, column=col_idx + 1, value=new_value)
        
        # 計算並更新總數 (不含總數欄本身)
        total = 0
        for label, idx in config:
            if label != '總數':
                cell_val = ws.cell(row=target_row, column=idx + 1).value
                total += int(cell_val) if cell_val else 0
        
        # 找出總數欄索引
        total_col_idx = None
        for label, idx in config:
            if label == '總數':
                total_col_idx = idx
                break
        
        if total_col_idx is not None:
            ws.cell(row=target_row, column=total_col_idx + 1, value=total)
        
        wb.save(casting_file)
        wb.close()
        
        # 記錄歷程
        log_edit(part_type, str(item_id) if item_id else model_name, field, old_value, new_value, user_id)
        
        return {'success': True, 'old_value': old_value, 'total': total, 'item_id': item_id}
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': str(e)}

def log_edit(part_type, item_id, field, old_value, new_value, user_id, note=None):
    """記錄編輯歷程"""
    try:
        log_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'logs')
        os.makedirs(log_dir, exist_ok=True)
        log_file = os.path.join(log_dir, 'edit_history.json')
        
        history = []
        if os.path.exists(log_file):
            with open(log_file, 'r', encoding='utf-8') as f:
                history = json.load(f)
        
        entry = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'user': user_id,
            'part': part_type,
            'item_id': item_id,
            'field': field,
            'old_value': old_value,
            'new_value': new_value
        }
        
        if note:
            entry['note'] = note
            
        history.insert(0, entry)
        
        # 只保留最近 500 筆
        history = history[:500]
        
        with open(log_file, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
    
    except Exception as e:
        print(f"Error logging edit: {e}")

def get_edit_history(part_type, limit=50):
    """取得特定鑄件的修改歷程"""
    try:
        log_file = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'logs', 'edit_history.json')
        
        if not os.path.exists(log_file):
            return []
        
        with open(log_file, 'r', encoding='utf-8') as f:
            history = json.load(f)
        
        # 過濾該鑄件的歷程
        filtered = [h for h in history if h.get('part') == part_type]
        return filtered[:limit]
    
    except Exception as e:
        print(f"Error reading history: {e}")
        return []

def get_item_history(part_type, item_id, limit=100):
    """取得特定品號的完整修改歷程"""
    try:
        log_file = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'logs', 'edit_history.json')
        
        if not os.path.exists(log_file):
            return []
        
        with open(log_file, 'r', encoding='utf-8') as f:
            history = json.load(f)
        
        # 過濾該品號的歷程
        filtered = [h for h in history if h.get('part') == part_type and h.get('item_id') == item_id]
        return filtered[:limit]
    
    except Exception as e:
        print(f"Error reading item history: {e}")
        return []

def get_history_stats(part_type):
    """取得歷程統計資訊"""
    try:
        log_file = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'logs', 'edit_history.json')
        
        if not os.path.exists(log_file):
            return {'total_edits': 0, 'unique_items': 0, 'recent_activity': []}
        
        with open(log_file, 'r', encoding='utf-8') as f:
            history = json.load(f)
        
        # 過濾該鑄件的歷程
        filtered = [h for h in history if h.get('part') == part_type]
        
        # 統計資訊
        unique_items = set(h.get('item_id') for h in filtered)
        unique_users = set(h.get('user') for h in filtered)
        
        # 最近活動（最近7天）
        from datetime import datetime, timedelta
        now = datetime.now()
        week_ago = now - timedelta(days=7)
        
        recent = []
        for h in filtered[:20]:  # 最近20筆
            try:
                ts = datetime.strptime(h.get('timestamp', ''), '%Y-%m-%d %H:%M:%S')
                if ts >= week_ago:
                    recent.append(h)
            except:
                pass
        
        return {
            'total_edits': len(filtered),
            'unique_items': len(unique_items),
            'unique_users': len(unique_users),
            'recent_activity': recent
        }
    
    except Exception as e:
        print(f"Error getting history stats: {e}")
        return {'total_edits': 0, 'unique_items': 0, 'recent_activity': []}


def get_stock_history(operation_type=None, part_type=None, date_from=None, date_to=None, keyword=None, limit=200):
    """查詢入庫/出庫歷史記錄
    
    Args:
        operation_type: 'in' (入庫), 'out' (出庫), or None (全部)
        part_type: 零件類型 (底座/工作台/橫樑/立柱), or None (全部)
        date_from: 開始日期 (YYYY-MM-DD)
        date_to: 結束日期 (YYYY-MM-DD)
        keyword: 關鍵字搜尋 (品號、備註等)
        limit: 最大回傳筆數
    """
    try:
        log_file = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'logs', 'edit_history.json')
        
        if not os.path.exists(log_file):
            return {'records': [], 'stats': {'total': 0, 'stock_in': 0, 'stock_out': 0}}
        
        with open(log_file, 'r', encoding='utf-8') as f:
            history = json.load(f)
        
        records = []
        stats = {'total': 0, 'stock_in': 0, 'stock_out': 0}
        

        # 加裝機型查找表
        _ITEM_MODEL_MAP = {}
        try:
            casting_file = current_app.config['CASTING_FILE']
            import pandas as pd
            xl = pd.ExcelFile(casting_file)
            for part_name, sheet_idx in [('底座', 1), ('工作台', 2), ('橫樑', 3), ('立柱', 4)]:
                df = pd.read_excel(xl, sheet_name=sheet_idx)
                for _, row in df.iterrows():
                    p_id = str(row.iloc[0]).strip()
                    model = str(row.iloc[1]).strip()
                    if p_id and p_id != 'nan' and p_id != '品號' and model and model != 'nan':
                        _ITEM_MODEL_MAP[p_id] = model
        except Exception as ex:
            print(f"Error building model map: {ex}")

        for h in history:
            field = h.get('field', '')
            note_text = h.get('note', '')
            
            # 判斷是否為入庫或出庫記錄
            is_stock_in = '入庫' in field
            is_stock_out = '出庫' in field
            
            if not is_stock_in and not is_stock_out:
                continue
            
            # 過濾操作類型
            if operation_type == 'in' and not is_stock_in:
                continue
            if operation_type == 'out' and not is_stock_out:
                continue
            
            # 過濾零件類型
            if part_type and h.get('part') != part_type:
                continue
            
            # 過濾日期範圍
            if date_from or date_to:
                try:
                    ts = datetime.strptime(h.get('timestamp', ''), '%Y-%m-%d %H:%M:%S')
                    if date_from:
                        from_dt = datetime.strptime(date_from, '%Y-%m-%d')
                        if ts < from_dt:
                            continue
                    if date_to:
                        to_dt = datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                        if ts > to_dt:
                            continue
                except:
                    pass
            
            # 關鍵字搜尋 (品號、備註、field)
            if keyword:
                kw = keyword.lower()
                searchable = ' '.join([
                    str(h.get('item_id', '')),
                    str(h.get('note', '')),
                    str(h.get('field', '')),
                    str(h.get('user', '')),
                    str(h.get('part', ''))
                ]).lower()
                if kw not in searchable:
                    continue
            
            # 解析工單編碼 (從 note 或 field 中解析)
            order_code = ''
            if note_text:
                import re
                # 優先抓取「工單編碼」
                oc_match = re.search(r'工單編碼:\s*([^,]+)', str(note_text))
                if oc_match:
                    order_code = oc_match.group(1).strip()
                else:
                    # 備案：抓取「鑄件編號」
                    cn_match = re.search(r'鑄件編號:\s*([^,]+)', str(note_text))
                    if cn_match:
                        order_code = cn_match.group(1).strip()
            
            record = {
                'timestamp': h.get('timestamp', ''),
                'user': h.get('user', ''),
                'part': h.get('part', ''),
                'item_id': h.get('item_id', ''),
                'model': _ITEM_MODEL_MAP.get(str(h.get('item_id', '')), ''),
                'field': field,
                'old_value': h.get('old_value', 0),
                'new_value': h.get('new_value', 0),
                'note': note_text,
                'operation': '入庫' if is_stock_in else '出庫',
                'quantity': abs((h.get('new_value') or 0) - (h.get('old_value') or 0)),
                'order_code': order_code
            }
            
            # 從 field 或 note 中解析供應商/工單資訊
            if is_stock_in:
                # e.g. "素材 (入庫: 正鋒(CFC))"
                import re
                supplier_match = re.search(r'入庫:\s*(.+)\)', field)
                record['supplier'] = supplier_match.group(1).strip() if supplier_match else ''
                
                # 如果 field 沒抓到，從 note 抓
                if not record['supplier'] and note_text:
                    s_note_match = re.search(r'鑄造商:\s*([^,]+)', str(note_text))
                    if s_note_match:
                        record['supplier'] = s_note_match.group(1).strip()
            elif is_stock_out:
                # e.g. "成品研磨 (出庫: 工單 123456)"
                import re
                wo_match = re.search(r'工單\s*(\S+)', field)
                po_match = re.search(r'採購單\s*(\S+)', field)
                
                parsed_wo = wo_match.group(1).rstrip(',)') if wo_match else ''
                record['work_order'] = parsed_wo
                record['purchase_order'] = po_match.group(1).rstrip(')') if po_match else ''
                
                # 從 note 中特別抓取鑄造商 (如果是出庫時有特別標記)
                if note_text:
                    s_note_match = re.search(r'鑄造商:\s*([^,]+)', str(note_text))
                    if s_note_match:
                        record['supplier'] = s_note_match.group(1).strip()
                
                # 如果 note 沒寫工單，用 field 裡的
                if not record['order_code'] and parsed_wo:
                    record['order_code'] = parsed_wo
            
            records.append(record)
            
            # 更新統計
            if is_stock_in:
                stats['stock_in'] += 1
            if is_stock_out:
                stats['stock_out'] += 1
        
        stats['total'] = len(records)
        
        return {
            'records': records[:limit],
            'stats': stats
        }
    
    except Exception as e:
        print(f"Error getting stock history: {e}")
        import traceback
        traceback.print_exc()
        return {'records': [], 'stats': {'total': 0, 'stock_in': 0, 'stock_out': 0}}



def stock_in_material(part_name, quantity, model=None, supplier=None, user=None, work_order_code=None, barcode=None):
    """
    入庫操作 - 增加素材數量
    """
    try:
        from flask import current_app, session
        import pandas as pd
        
        # 嘗試在模組內引用 log_edit，如果不行的話可能需要重新組織代碼
        # 由於 log_edit 就在同一個模組，應該可以直接使用，但因為我們是覆寫，
        # 需要確保在函數內部能訪問到。
        # 為了安全起見，如果此函數在 inventory.py 底部，它應該能看到上面的 log_edit
        
        casting_file = current_app.config['CASTING_FILE']
        
        # 零件對應的工作表索引
        sheet_mapping = {
            '底座': 1,
            '工作台': 2,
            '橫樑': 3,
            '立柱': 4
        }
        
        if part_name not in sheet_mapping:
            return {'success': False, 'error': f'未知的零件名稱: {part_name}'}
        
        sheet_idx = sheet_mapping[part_name]
        
        # 動態決定素材欄位索引
        config = CONFIGS.get(part_name, [])
        material_col_idx = 2  # 預設值
        for label, idx in config:
            if label == '素材':
                material_col_idx = idx
                break
        
        # 讀取 Excel 檔案
        xl = pd.ExcelFile(casting_file)
        df = pd.read_excel(xl, sheet_name=sheet_idx)
        
        updated = False
        log_data = None
        
        # 根據品號 (model) 更新
        if model:
            search_part_no = str(model).strip()
            for idx, row in df.iterrows():
                # 第0欄是品號
                curr_part_no = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
                
                # 如果品號匹配
                if curr_part_no == search_part_no:
                    current_val = row.iloc[material_col_idx]
                    old_v = int(float(current_val)) if pd.notna(current_val) else 0
                    new_val = old_v + quantity
                    df.at[idx, df.columns[material_col_idx]] = new_val
                    updated = True
                    
                    log_data = {
                        'item_id': curr_part_no,
                        'old_value': old_v,
                        'new_value': new_val
                    }
                    break
        
        # 如果沒有指定品號，或找不到品號，則使用舊邏輯 (找第一個有效機型)
        if not updated and not model:
            for idx, row in df.iterrows():
                model_val = row.get('機型', '')
                if pd.notna(model_val) and str(model_val).strip() and str(model_val).strip() != '品號':
                    current_val = row.iloc[material_col_idx]
                    old_v = int(float(current_val)) if pd.notna(current_val) else 0
                    new_val = old_v + quantity
                    df.at[idx, df.columns[material_col_idx]] = new_val
                    updated = True

                    log_data = {
                        'item_id': str(model_val).strip(), # 如果沒有品號，用機型名暫代
                        'old_value': old_v,
                        'new_value': new_val
                    }
                    break
        
        if not updated:
            if model:
                return {'success': False, 'error': f'找不到指定的品號: {model}'}
            else:
                return {'success': False, 'error': '找不到有效的機型可以入庫'}
        
        # 寫回 Excel
        with pd.ExcelWriter(casting_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=xl.sheet_names[sheet_idx], index=False)
        
        # 寫入履歷
        if log_data:
            user_id = user if user else session.get('user', 'System')
            field_name = '素材'
            if supplier:
                field_name += f' (入庫: {supplier})'
            else:
                field_name += ' (入庫)'
                
            note = None
            note_parts = []
            if work_order_code:
                note_parts.append(f"工單編碼: {work_order_code}")
            if barcode:
                note_parts.append(f"鑄件編號: {barcode}")
            if note_parts:
                note = ', '.join(note_parts)
                
            try:
                log_edit(part_name, log_data['item_id'], field_name, 
                         log_data['old_value'], log_data['new_value'], user_id, note)
            except Exception as e:
                print(f"Failed to log edit: {e}")

        msg = f'{part_name} 入庫成功，已增加 {quantity} 件到素材'
        if model:
            msg += f' (品號: {model})'
        if supplier:
            msg += f'，供應商: {supplier}'
            
        return {
            'success': True,
            'message': msg,
            'quantity': quantity
        }
        
    except Exception as e:
        return {'success': False, 'error': str(e)}


def stock_out_product(part_name, work_order, quantity, model=None, purchase_order=None, supplier=None):
    """
    出庫操作 - 扣除成品數量
    """
    try:
        from flask import current_app, session
        import pandas as pd
        
        casting_file = current_app.config['CASTING_FILE']
        
        sheet_mapping = {
            '底座': 1,
            '工作台': 2,
            '橫樑': 3,
            '立柱': 4
        }
        
        if part_name not in sheet_mapping:
            return {'success': False, 'error': f'未知的零件名稱: {part_name}'}
        
        sheet_idx = sheet_mapping[part_name]
        
        # 動態決定成品欄位索引
        config = CONFIGS.get(part_name, [])
        product_col_idx = None
        product_label = '成品'
        for label, idx in config:
            if label in ['成品', '成品研磨']:
                product_col_idx = idx
                product_label = label
                break
        
        if product_col_idx is None:
            return {'success': False, 'error': f'找不到 {part_name} 的成品欄位'}

        xl = pd.ExcelFile(casting_file)
        df = pd.read_excel(xl, sheet_name=sheet_idx)
        
        updated = False
        log_data = None
        
        # 邏輯 A: 指定品號出庫
        if model:
            search_part_no = str(model).strip()
            target_idx = -1
            
            # 尋找目標行
            for idx, row in df.iterrows():
                curr_part_no = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ''
                if curr_part_no == search_part_no:
                    target_idx = idx
                    val = row.iloc[product_col_idx]
                    current_stock = int(float(val)) if pd.notna(val) else 0
                    break
            
            if target_idx == -1:
                return {'success': False, 'error': f'找不到指定的品號: {model}'}
            
            if current_stock < quantity:
                return {'success': False, 'error': f'庫存不足！品號 {model} 當前成品: {current_stock}，需要: {quantity}'}
            
            # 執行扣除
            new_val = current_stock - quantity
            df.at[target_idx, df.columns[product_col_idx]] = new_val
            updated = True
            
            log_data = {
                'item_id': search_part_no,
                'old_value': current_stock,
                'new_value': new_val
            }
            
        # 邏輯 B: 未指定品號
        else:
            # 省略舊邏輯的 log，因為現在都強制選機型了
            # 如果真的跑進來這裡，我們就只記錄操作成功，數字可能不太精確因為是扣多行
            
            # 計算總庫存
            current_total = 0
            for _, row in df.iterrows():
                m_val = row.get('機型', '')
                if pd.notna(m_val) and str(m_val).strip() and str(m_val).strip() != '品號':
                    val = row.iloc[product_col_idx]
                    if pd.notna(val):
                        try:
                            current_total += int(float(val))
                        except:
                            pass
            
            if current_total < quantity:
                return {'success': False, 'error': f'該零件總成品數量不足！當前: {current_total}，需要: {quantity}'}
            
            # 從成品中扣除數量
            remaining = quantity
            for idx, row in df.iterrows():
                if remaining <= 0:
                    break
                    
                m_val = row.get('機型', '')
                if pd.notna(m_val) and str(m_val).strip() and str(m_val).strip() != '品號':
                    current_val = row.iloc[product_col_idx]
                    if pd.notna(current_val):
                        try:
                            current_qty = int(float(current_val))
                            if current_qty > 0:
                                deduct = min(current_qty, remaining)
                                df.at[idx, df.columns[product_col_idx]] = current_qty - deduct
                                remaining -= deduct
                                updated = True
                        except:
                            pass
        
        if not updated:
             return {'success': False, 'error': '出庫操作未能完成，請檢查庫存狀態'}

        # 寫回 Excel
        with pd.ExcelWriter(casting_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=xl.sheet_names[sheet_idx], index=False)
            
        # 寫入履歷
        # 如果是邏輯A (有log_data)，我們記錄精確值
        if log_data:
            user_id = session.get('user', 'System')
            field_name = f'{product_label} (出庫: 工單 {work_order}'
            if purchase_order:
                field_name += f', 採購單 {purchase_order}'
            field_name += ')'
            
            # 將鑄造商資訊也存入 note
            note = f"工單編碼: {work_order}"
            if supplier:
                note += f", 鑄造商: {supplier}"
                
            try:
                log_edit(part_name, log_data['item_id'], field_name, 
                         log_data['old_value'], log_data['new_value'], user_id, note)
            except Exception as e:
                print(f"Failed to log edit: {e}")
        
        # 如果是邏輯B，我們目前沒辦法精確記錄單一 item_id 的變化，除非記錄多筆
        # 暫時不處理邏輯B的 Log，因為前端已經強制要選機型了。
        
        return {
            'success': True,
            'message': f'{part_name} 出庫成功，工單 {work_order}，已扣除 {quantity} 件成品',
            'work_order': work_order,
            'purchase_order': purchase_order,
            'quantity': quantity
        }
        
    except Exception as e:
        return {'success': False, 'error': str(e)}
