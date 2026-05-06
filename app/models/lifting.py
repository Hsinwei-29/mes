import pandas as pd
from flask import current_app
from openpyxl import load_workbook
from datetime import datetime
import json
import os
import glob

# 快取
_LIFTING_CACHE = {'mtime': 0, 'data': None}

def get_lifting_file_path():
    """尋找吊具清冊的路徑"""
    # 指定目前已確認位置的檔案
    target_path = '/home/hsinwei/mes/吊具吊鍊清冊.xlsx'
    if os.path.exists(target_path):
        return target_path
    
    # 原有的搜尋邏輯做為備援
    files = glob.glob('D:\\app\\mes\\*吊具*.xlsx')
    if files:
        return files[0]
    return os.path.join(os.path.dirname(__file__), '..', '..', '吊具清冊.xlsx')

SHEET_NAMES = ['吊具', '吊鍊', '布帶']


def ensure_columns(df):
    """確保有借用人跟借用日期欄位"""
    if '目前借用人' not in df.columns:
        df['目前借用人'] = ''
    if '借用日期' not in df.columns:
        df['借用日期'] = ''
    return df

def load_lifting_inventory():
    """載入吊具清單"""
    file_path = get_lifting_file_path()
    if not os.path.exists(file_path):
        return {}

    current_mtime = os.path.getmtime(file_path)
    if _LIFTING_CACHE['mtime'] == current_mtime and _LIFTING_CACHE['data']:
        return _LIFTING_CACHE['data']

    # 嘗試讀取持久化快取
    cache_dir = os.path.join(os.getcwd(), 'app', 'cache')
    os.makedirs(cache_dir, exist_ok=True)
    cache_file = os.path.join(cache_dir, 'lifting_cache.pkl')
    if os.path.exists(cache_file):
        try:
            import pickle
            with open(cache_file, 'rb') as f:
                cached = pickle.load(f)
            if cached.get('mtime') == current_mtime:
                _LIFTING_CACHE['mtime'] = current_mtime
                _LIFTING_CACHE['data'] = cached['data']
                return cached['data']
        except:
            pass

    try:
        xl = pd.ExcelFile(file_path, engine='openpyxl')
        data = {}
        
        # 輔助：放置位置排序關鍵值
        def get_location_sort_val(loc_str):
            loc = str(loc_str).strip()
            # 正規化：去除換行、空白，統一「第1區」與「第一區」
            loc = loc.replace('\n', '').replace(' ', '')
            
            mapping = {
                '第一區': 1, '第1區': 1,
                '第二區': 2, '第2區': 2,
                '第三區': 3, '第3區': 3,
                '第四區': 4, '第4區': 4,
                '第五區': 5, '第5區': 5,
                '第六區': 6, '第6區': 6,
                '第七區': 7, '第7區': 7,
                '第八區': 8, '第8區': 8,
                '第九區': 9, '第9區': 9,
                '第十區': 10, '第10區': 10
            }
            # 如果符合對照表則回傳對應數字，否則回傳一個極大值
            return mapping.get(loc, 999), loc

        for sheet in SHEET_NAMES:
            if sheet in xl.sheet_names:
                df = pd.read_excel(xl, sheet_name=sheet)
                df = df.fillna('')
                df = ensure_columns(df)
                
                # Get name map
                from .user import User
                name_map = User.get_name_map()
                
                # Convert DataFrame to list of dicts
                items = []
                for _, row in df.iterrows():
                    # 吊具編號是唯一識別
                    # 吊具編號是唯一識別
                    item_id = str(row.get('吊具編號', '')).strip()
                    if item_id.endswith('.0'):
                        item_id = item_id[:-2]
                    if not item_id or item_id in ['nan', 'N/A', 'None']:
                        continue
                    
                    loc_val = str(row.get('放置位置', '')).strip()
                    replace_map = {
                        '第一區': '第1區', '第二區': '第2區', '第三區': '第3區', '第四區': '第4區', '第五區': '第5區',
                        '第六區': '第6區', '第七區': '第7區', '第八區': '第8區', '第九區': '第9區', '第十區': '第10區'
                    }
                    for k, v in replace_map.items():
                        loc_val = loc_val.replace(k, v)
                        
                    raw_borrower = str(row.get('目前借用人', '')).strip()
                    borrower_display = name_map.get(raw_borrower, raw_borrower)

                    items.append({
                        'category': sheet,
                        'id': item_id,
                        'spec': str(row.get('吊具規格 /重量/長度', '')),
                        'location': loc_val,
                        'status': str(row.get('使用狀態', '')),
                        'borrower': borrower_display,
                        'borrower_raw': raw_borrower,
                        'borrow_date': str(row.get('借用日期', ''))
                    })
                
                # 按是否借用中 (借用中排在最前面)，再按放置位置 (第1區、第2區...) 排序
                items.sort(key=lambda x: (
                    0 if x['status'] == '借用中' or x['borrower_raw'] else 1,
                    get_location_sort_val(x['location'])[0]
                ))
                
                data[sheet] = items
                
        _LIFTING_CACHE['mtime'] = current_mtime
        _LIFTING_CACHE['data'] = data
        
        try:
            import pickle
            with open(cache_file, 'wb') as f:
                pickle.dump({'mtime': current_mtime, 'data': data}, f)
        except:
            pass
            
        return data
    except Exception as e:
        print(f"Error loading lifting inventory: {e}")
        return {}


def log_lifting_action(category, item_id, action, user_name):
    """紀錄吊具操作歷史"""
    history_file = '/home/hsinwei/app/mes/logs/lifting_history.json'
    history = []
    if os.path.exists(history_file):
        try:
            with open(history_file, 'r', encoding='utf-8') as f:
                history = json.load(f)
        except:
            history = []
            
    from .user import User
    name_map = User.get_name_map()
    user_display = name_map.get(user_name, user_name)
    
    record = {
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'category': category,
        'item_id': item_id,
        'action': '領用' if action == 'borrow' else '歸還',
        'user': user_display
    }
    history.insert(0, record)
    # 限制紀錄數量
    history = history[:1000]
    
    os.makedirs(os.path.dirname(history_file), exist_ok=True)
    with open(history_file, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)

def get_lifting_history():
    """取得吊具操作歷史"""
    history_file = '/home/hsinwei/app/mes/logs/lifting_history.json'
    if os.path.exists(history_file):
        try:
            with open(history_file, 'r', encoding='utf-8') as f:
                history = json.load(f)
                
            # 確保舊紀錄中的英文名字也能對應到目前的中文名 (選用，但較貼心)
            from .user import User
            name_map = User.get_name_map()
            for h in history:
                if h.get('user') in name_map:
                    h['user'] = name_map[h['user']]
            return history
        except:
            return []
    return []


def update_lifting_status(category, item_id, action, user_name):
    """
    更新吊具狀態
    action: 'borrow' 或 'return'
    """
    file_path = get_lifting_file_path()
    if not os.path.exists(file_path):
        return False, "找不到吊具清冊檔案"

    try:
        wb = load_workbook(file_path)
        if category not in wb.sheetnames:
            return False, f"找不到工作表: {category}"

        ws = wb[category]
        headers = {cell.value: idx for idx, cell in enumerate(ws[1])}
        
        # 確保新的兩欄存在
        if '目前借用人' not in headers:
            new_col_idx = len(headers) + 1
            ws.cell(row=1, column=new_col_idx, value='目前借用人')
            headers['目前借用人'] = new_col_idx - 1
            
        if '借用日期' not in headers:
            new_col_idx = len(headers) + 1
            ws.cell(row=1, column=new_col_idx, value='借用日期')
            headers['借用日期'] = new_col_idx - 1

        if '吊具編號' not in headers or '使用狀態' not in headers:
            return False, "工作表缺少必要標題：'吊具編號' 或 '使用狀態'"

        id_col = headers['吊具編號']
        status_col = headers['使用狀態']
        borrower_col = headers['目前借用人']
        date_col = headers['借用日期']

        row_found = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            curr_id = str(row[id_col]).strip() if row[id_col] is not None else ''
            if curr_id.endswith('.0'):
                curr_id = curr_id[:-2]
            if curr_id == str(item_id):
                row_found = row_idx
                break

        if not row_found:
            return False, f"找不到吊具編號: {item_id}"

        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        if action == 'borrow':
            ws.cell(row=row_found, column=status_col + 1, value='借用中')
            ws.cell(row=row_found, column=borrower_col + 1, value=user_name)
            ws.cell(row=row_found, column=date_col + 1, value=now_str)
        elif action == 'return':
            ws.cell(row=row_found, column=status_col + 1, value='在庫')
            ws.cell(row=row_found, column=borrower_col + 1, value='')
            ws.cell(row=row_found, column=date_col + 1, value='')
        else:
            return False, "未知的操作"

        wb.save(file_path)
        
        # 紀錄歷程
        log_lifting_action(category, item_id, action, user_name)
        
        # 強制刷新快取
        _LIFTING_CACHE['mtime'] = 0 
        return True, "更新成功"

    except PermissionError:
        return False, "檔案可能被另一個程式(如 Excel)開啟中，請先關閉。"
    except Exception as e:
        print(f"Error updating lifting status: {e}")
        return False, str(e)
