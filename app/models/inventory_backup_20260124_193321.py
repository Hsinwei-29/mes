import pandas as pd
from flask import current_app
from openpyxl import load_workbook
from datetime import datetime
import json
import os

def load_casting_inventory():
    """載入鑄件庫存資料"""
    try:
        casting_file = current_app.config['CASTING_FILE']
        xl = pd.ExcelFile(casting_file)
        
        # 從各個鑄件工作表計算總數
        inventory = {}
        for part_name, sheet_idx in [('底座', 1), ('工作台', 2), ('橫樑', 3), ('立柱', 4)]:
            try:
                df = pd.read_excel(xl, sheet_name=sheet_idx)
                numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
                # 對所有數字列求和（排除第一列品號）
                total = 0
                if len(numeric_cols) > 1:
                    for col in numeric_cols[1:]:  # 跳過第一個數字列（品號）
                        col_sum = df[col].sum()
                        if not pd.isna(col_sum):
                            total += col_sum
                inventory[part_name] = int(total) if total > 0 else 0
            except Exception as e:
                print(f"Error loading {part_name}: {e}")
                inventory[part_name] = 0
        
        # 重新排序
        inventory = {'工作台': inventory.get('工作台', 0), '底座': inventory.get('底座', 0), 
                     '橫樑': inventory.get('橫樑', 0), '立柱': inventory.get('立柱', 0)}
        
        # 原始代碼保留用於詳細資料
        df = pd.read_excel(xl, sheet_name=2)  # 使用工作台工作表
        
        # 取得庫存總計 (調整順序：工作台在底座之前)
        inventory = {
            '工作台': df['工作台'].sum() if '工作台' in df.columns else 0,
            '底座': df['底座'].sum() if '底座' in df.columns else 0,
            '橫樑': df['橫樑'].sum() if '橫樑' in df.columns else 0,
            '立柱': df['立柱'].sum() if '立柱' in df.columns else 0,
        }
        
        # 轉換 NaN 為 0
        for key in inventory:
            if pd.isna(inventory[key]):
                inventory[key] = 0
            else:
                inventory[key] = int(inventory[key])
        
        # 按機型整理詳細資料
        details = []
        for _, row in df.iterrows():
            model = row['機型'] if '機型' in df.columns else row.iloc[0]
            detail = {'機型': str(model)}
            for col in ['底座', '工作台', '橫樑', '立柱']:
                if col in df.columns:
                    val = row[col]
                    detail[col] = 0 if pd.isna(val) else int(val)
                else:
                    detail[col] = 0
            # 只加入有庫存的機型
            if any(detail[col] > 0 for col in ['底座', '工作台', '橫樑', '立柱']):
                details.append(detail)
        
        return {'summary': inventory, 'details': details}
    except Exception as e:
        print(f"Error loading casting inventory: {e}")
        return {'summary': {}, 'details': [], 'error': str(e)}

# 欄位配置 (全域，供多個函式使用)
CONFIGS = {
    '底座': [('素材', 2), ('製程四', 3), ('製程三', 4), ('成品研磨', 5), ('總數', 7)],
    '工作台': [('素材', 2), ('製程一', 3), ('製程二', 4), ('製程三', 5), ('製程四', 6), ('成品', 7), ('總數', 9)],
    '橫樑': [('素材', 2), ('製程六', 4), ('製程五', 5), ('成品研磨', 6), ('總數', 8)],
    '立柱': [('素材', 2), ('半品', 3), ('成品銑工', 4), ('成品研磨', 5), ('總數', 6)]
}

SHEET_MAP = {'底座': 1, '工作台': 2, '橫樑': 3, '立柱': 4}

def get_part_details(part_type):
    """載入特定鑄件的詳細製程資料 (動態表頭)"""
    try:
        casting_file = current_app.config['CASTING_FILE']
        sheet_idx = SHEET_MAP.get(part_type)
        if sheet_idx is None: return {"headers": [], "rows": []}

        config = CONFIGS.get(part_type, [])
        headers = ['品號', '機型'] + [c[0] for c in config]
        
        df = pd.read_excel(casting_file, sheet_name=sheet_idx)
        
        rows = []
        for _, row in df.iterrows():
            if pd.isna(row.iloc[0]) and pd.isna(row.iloc[1]): continue
            
            def to_int(val):
                try: return int(float(val)) if pd.notna(val) else 0
                except: return 0

            # 建立資料列
            data_row = {
                '品號': str(row.iloc[0]),
                '機型': str(row.iloc[1])
            }
            # 動態加入各階段數據
            has_data = False
            for label, idx in config:
                val = to_int(row.iloc[idx])
                data_row[label] = val
                if val > 0: has_data = True
            
            if has_data:
                rows.append(data_row)
                
        return {"headers": headers, "rows": rows}
    except Exception as e:
        print(f"Error loading {part_type} details: {e}")
        return {"headers": [], "rows": []}

def update_cell(part_type, item_id, field, new_value, user_id):
    """更新 Excel 儲存格並記錄歷程"""
    try:
        casting_file = current_app.config['CASTING_FILE']
        sheet_idx = SHEET_MAP.get(part_type)
        config = CONFIGS.get(part_type, [])
        
        if sheet_idx is None:
            return {'success': False, 'error': '無效的鑄件類型'}
        
        # 找出欄位索引
        col_idx = None
        for label, idx in config:
            if label == field:
                col_idx = idx
                break
        
        if col_idx is None:
            return {'success': False, 'error': '無效的欄位'}
        
        # 使用 openpyxl 開啟並修改
        wb = load_workbook(casting_file)
        ws = wb.worksheets[sheet_idx]
        
        # 找出對應行 (品號在第一欄)
        target_row = None
        old_value = None
        for row_num in range(2, ws.max_row + 1):  # 跳過標題行
            cell_value = ws.cell(row=row_num, column=1).value
            if str(cell_value) == str(item_id):
                target_row = row_num
                old_value = ws.cell(row=row_num, column=col_idx + 1).value  # openpyxl 1-indexed
                break
        
        if target_row is None:
            return {'success': False, 'error': '找不到品號'}
        
        # 更新儲存格 (openpyxl column is 1-indexed)
        ws.cell(row=target_row, column=col_idx + 1, value=new_value)
        
        # 計算並更新總數 (不含總數欄本身)
        total = 0
        for label, idx in config:
            if label != '總數':
                cell_val = ws.cell(row=target_row, column=idx + 1).value
                total += int(cell_val) if cell_val else 0
        
        # 找出總數欄索引
        total_col_idx = None
        for label, idx in config:
            if label == '總數':
                total_col_idx = idx
                break
        
        if total_col_idx:
            ws.cell(row=target_row, column=total_col_idx + 1, value=total)
        
        wb.save(casting_file)
        wb.close()
        
        # 記錄歷程
        log_edit(part_type, item_id, field, old_value, new_value, user_id)
        
        return {'success': True, 'old_value': old_value, 'total': total}
    
    except Exception as e:
        print(f"Error updating cell: {e}")
        return {'success': False, 'error': str(e)}

def log_edit(part_type, item_id, field, old_value, new_value, user_id):
    """記錄編輯歷程"""
    try:
        log_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'logs')
        os.makedirs(log_dir, exist_ok=True)
        log_file = os.path.join(log_dir, 'edit_history.json')
        
        history = []
        if os.path.exists(log_file):
            with open(log_file, 'r', encoding='utf-8') as f:
                history = json.load(f)
        
        history.insert(0, {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'user': user_id,
            'part': part_type,
            'item_id': item_id,
            'field': field,
            'old_value': old_value,
            'new_value': new_value
        })
        
        # 只保留最近 500 筆
        history = history[:500]
        
        with open(log_file, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
    
    except Exception as e:
        print(f"Error logging edit: {e}")

def get_edit_history(part_type, limit=50):
    """取得特定鑄件的修改歷程"""
    try:
        log_file = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'logs', 'edit_history.json')
        
        if not os.path.exists(log_file):
            return []
        
        with open(log_file, 'r', encoding='utf-8') as f:
            history = json.load(f)
        
        # 過濾該鑄件的歷程
        filtered = [h for h in history if h.get('part') == part_type]
        return filtered[:limit]
    
    except Exception as e:
        print(f"Error reading history: {e}")
        return []

def get_item_history(part_type, item_id, limit=100):
    """取得特定品號的完整修改歷程"""
    try:
        log_file = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'logs', 'edit_history.json')
        
        if not os.path.exists(log_file):
            return []
        
        with open(log_file, 'r', encoding='utf-8') as f:
            history = json.load(f)
        
        # 過濾該品號的歷程
        filtered = [h for h in history if h.get('part') == part_type and h.get('item_id') == item_id]
        return filtered[:limit]
    
    except Exception as e:
        print(f"Error reading item history: {e}")
        return []

def get_history_stats(part_type):
    """取得歷程統計資訊"""
    try:
        log_file = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'logs', 'edit_history.json')
        
        if not os.path.exists(log_file):
            return {'total_edits': 0, 'unique_items': 0, 'recent_activity': []}
        
        with open(log_file, 'r', encoding='utf-8') as f:
            history = json.load(f)
        
        # 過濾該鑄件的歷程
        filtered = [h for h in history if h.get('part') == part_type]
        
        # 統計資訊
        unique_items = set(h.get('item_id') for h in filtered)
        unique_users = set(h.get('user') for h in filtered)
        
        # 最近活動（最近7天）
        from datetime import datetime, timedelta
        now = datetime.now()
        week_ago = now - timedelta(days=7)
        
        recent = []
        for h in filtered[:20]:  # 最近20筆
            try:
                ts = datetime.strptime(h.get('timestamp', ''), '%Y-%m-%d %H:%M:%S')
                if ts >= week_ago:
                    recent.append(h)
            except:
                pass
        
        return {
            'total_edits': len(filtered),
            'unique_items': len(unique_items),
            'unique_users': len(unique_users),
            'recent_activity': recent
        }
    
    except Exception as e:
        print(f"Error getting history stats: {e}")
        return {'total_edits': 0, 'unique_items': 0, 'recent_activity': []}
