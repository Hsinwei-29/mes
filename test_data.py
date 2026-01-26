import pandas as pd
import openpyxl

# 測試讀取 Excel
wb = openpyxl.load_workbook('鑄件盤點資料.xlsx')
print("所有工作表:", wb.sheetnames)
print()

# 測試每個工作表
for idx in range(len(wb.sheetnames)):
    print(f"=== 工作表 {idx}: {wb.sheetnames[idx]} ===")
    try:
        df = pd.read_excel('鑄件盤點資料.xlsx', sheet_name=idx)
        print(f"形狀: {df.shape}")
        print(f"列名: {df.columns.tolist()[:10]}")
        
        # 找出數字列
        numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        print(f"數字列: {numeric_cols[:5]}")
        
        if len(numeric_cols) > 0:
            last_num_col = numeric_cols[-1]
            total = df[last_num_col].sum()
            print(f"最後數字列 '{last_num_col}' 總和: {total}")
        
        print()
    except Exception as e:
        print(f"錯誤: {e}\n")
