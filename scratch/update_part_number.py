from openpyxl import load_workbook

file_path = '/home/hsinwei/app/mes/鑄件盤點資料.xlsx'
wb = load_workbook(file_path)
ws = wb['立柱']

target_model = 'LG-1370-NEO+200H'
part_number = '8811001625'

found = False
for row in range(2, ws.max_row + 1):
    model_name = str(ws.cell(row=row, column=2).value).strip()
    if model_name == target_model:
        ws.cell(row=row, column=1, value=part_number)
        print(f"Updated part number for {target_model} to {part_number} at row {row}")
        found = True
        break

if not found:
    print(f"Could not find model {target_model} in 立柱 sheet")

wb.save(file_path)
print("Excel file updated.")
