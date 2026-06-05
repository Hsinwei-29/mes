from openpyxl import load_workbook

file_path = '/home/hsinwei/app/mes/鑄件盤點資料.xlsx'
wb = load_workbook(file_path)

ws = wb['立柱']

# Print headers
headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
print("Headers:", headers)

# Print rows around row 6
for row in range(max(1, 6-3), min(ws.max_row+1, 6+4)):
    row_vals = [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"Row {row}: {row_vals}")
