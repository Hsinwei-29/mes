import os
import shutil
import glob
from openpyxl import load_workbook

# Paths
excel_file = '/home/hsinwei/app/mes/鑄件盤點資料.xlsx'
backup_dir = '/home/hsinwei/app/mes/workorder_backup'
cache_dir = '/home/hsinwei/app/mes/app/cache'

# 1. Backup the original file
os.makedirs(backup_dir, exist_ok=True)
backup_path = os.path.join(backup_dir, '鑄件盤資料_backup_before_lg800_column_update.xlsx')
shutil.copy2(excel_file, backup_path)
print(f"Backed up original file to: {backup_path}")

# 2. Update the Excel file
wb = load_workbook(excel_file)
sheet_name = '立柱'
if sheet_name not in wb.sheetnames:
    print(f"Error: Sheet '{sheet_name}' not found!")
    exit(1)

ws = wb[sheet_name]
target_model = 'LG-800'
new_part_number = '8811001113'
found = False

for row in range(2, ws.max_row + 1):
    model_val = ws.cell(row=row, column=2).value
    if model_val is not None:
        model_str = str(model_val).strip()
        if model_str.upper() == target_model.upper():
            old_part_number = ws.cell(row=row, column=1).value
            ws.cell(row=row, column=1, value=new_part_number)
            print(f"Updated row {row} in '{sheet_name}': Model='{model_val}', Old Part Number={old_part_number} -> New Part Number={new_part_number}")
            found = True
            break

if not found:
    print(f"Error: Could not find model '{target_model}' in sheet '{sheet_name}'!")
    exit(1)

wb.save(excel_file)
print("Excel file successfully updated and saved.")

# 3. Clear pickle caches
cache_files = glob.glob(os.path.join(cache_dir, '*.pkl'))
for f in cache_files:
    try:
        os.remove(f)
        print(f"Cleared cache file: {os.path.basename(f)}")
    except Exception as e:
        print(f"Error clearing cache file {f}: {e}")

# 4. Verify change by reloading
print("\n--- Verification ---")
wb_verify = load_workbook(excel_file)
ws_verify = wb_verify[sheet_name]
for row in range(2, ws_verify.max_row + 1):
    model_val = ws_verify.cell(row=row, column=2).value
    if model_val is not None and str(model_val).strip().upper() == target_model.upper():
        print(f"Verified row {row} in Excel: Model='{model_val}', Part Number={ws_verify.cell(row=row, column=1).value}")
        break
