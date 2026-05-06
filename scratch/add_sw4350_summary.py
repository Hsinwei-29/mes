from openpyxl import load_workbook

file_path = '/home/hsinwei/app/mes/鑄件盤點資料.xlsx'
wb = load_workbook(file_path)

ws_summary = wb['總數']
model_col = 1
for col in range(1, ws_summary.max_column + 1):
    if ws_summary.cell(row=1, column=col).value == '機型':
        model_col = col
        break

found = False
for row in range(2, ws_summary.max_row + 1):
    if str(ws_summary.cell(row=row, column=model_col).value).strip().upper() == 'SW4350':
        found = True
        break
if not found:
    new_row = ws_summary.max_row + 1
    ws_summary.cell(row=new_row, column=model_col, value='SW4350')
    print("Added SW4350 to 總數")

wb.save(file_path)
