from openpyxl import load_workbook

file_path = '/home/hsinwei/app/mes/鑄件盤點資料.xlsx'
wb = load_workbook(file_path)

# Update 工作台 sheet
ws_workbench = wb['工作台']
# Find if SW4320 exists
found = False
for row in range(2, ws_workbench.max_row + 1):
    if ws_workbench.cell(row=row, column=2).value == 'SW4320':
        found = True
        break
if not found:
    new_row = ws_workbench.max_row + 1
    ws_workbench.cell(row=new_row, column=1, value='N/A')
    ws_workbench.cell(row=new_row, column=2, value='SW4320')
    # Initialize other columns with 0
    for col in range(3, ws_workbench.max_column + 1):
        ws_workbench.cell(row=new_row, column=col, value=0)
    print("Added SW4320 to 工作台")

# Update 立柱 sheet
ws_column = wb['立柱']
# Find if LG1370neo+200H exists
found = False
for row in range(2, ws_column.max_row + 1):
    if str(ws_column.cell(row=row, column=2).value).strip().upper() == 'LG1370NEO+200H':
        found = True
        break
if not found:
    new_row = ws_column.max_row + 1
    ws_column.cell(row=new_row, column=1, value='N/A')
    ws_column.cell(row=new_row, column=2, value='LG1370neo+200H')
    # Initialize other columns with 0
    for col in range(3, ws_column.max_column + 1):
        ws_column.cell(row=new_row, column=col, value=0)
    print("Added LG1370neo+200H to 立柱")

# Update 總數 sheet
ws_summary = wb['總數']
# Summary sheet structure: Column 1 (A) is usually Model
# Let's check where '機型' is
model_col = 1
for col in range(1, ws_summary.max_column + 1):
    if ws_summary.cell(row=1, column=col).value == '機型':
        model_col = col
        break

for model in ['SW4320', 'LG1370neo+200H']:
    found = False
    for row in range(2, ws_summary.max_row + 1):
        if str(ws_summary.cell(row=row, column=model_col).value).strip().upper() == model.strip().upper():
            found = True
            break
    if not found:
        new_row = ws_summary.max_row + 1
        ws_summary.cell(row=new_row, column=model_col, value=model)
        print(f"Added {model} to 總數")

wb.save(file_path)
print("Excel file updated.")
