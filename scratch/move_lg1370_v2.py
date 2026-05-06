from openpyxl import load_workbook

file_path = '/home/hsinwei/app/mes/鑄件盤點資料.xlsx'
wb = load_workbook(file_path)

def move_model_after(sheet_name, target_model, after_model, model_col_idx):
    ws = wb[sheet_name]
    
    # 1. Collect all rows
    all_rows = []
    target_row_data = None
    
    for row in range(2, ws.max_row + 1):
        row_data = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        model_name = str(row_data[model_col_idx]).strip() if len(row_data) > model_col_idx else ""
        if model_name == target_model:
            target_row_data = row_data
            continue
        if any(row_data): # Keep non-empty rows
            all_rows.append(row_data)
    
    if not target_row_data:
        print(f"Target model {target_model} not found in {sheet_name}")
        return
    
    # 2. Find insertion point
    insert_idx = len(all_rows)
    for i, row_data in enumerate(all_rows):
        if str(row_data[model_col_idx]).strip() == after_model:
            insert_idx = i + 1
            break
            
    all_rows.insert(insert_idx, target_row_data)
    
    # 3. Write back
    for row in range(ws.max_row, 1, -1):
        ws.delete_rows(row)
        
    for row_data in all_rows:
        ws.append(row_data)
    
    print(f"Moved {target_model} after {after_model} in {sheet_name}")

# Move in 立柱 (Model is index 1)
# after LG-1370-NEO
move_model_after('立柱', 'LG-1370-NEO+200H', 'LG-1370-NEO', 1)

# Move in 總數 (Model is index 0)
# after LG-1370_NEO
move_model_after('總數', 'LG-1370_NEO+200H', 'LG-1370_NEO', 0)

wb.save(file_path)
print("Positions updated.")
