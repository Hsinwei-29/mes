from openpyxl import load_workbook

file_path = '/home/hsinwei/app/mes/鑄件盤點資料.xlsx'
wb = load_workbook(file_path)

def move_or_add_model(sheet_name, target_model, after_model, remove_names):
    ws = wb[sheet_name]
    
    # 1. Remove existing matches (including typos)
    rows_data = []
    headers = [cell.value for cell in ws[1]]
    
    # Collect all rows except headers
    all_rows = []
    for row in range(2, ws.max_row + 1):
        row_data = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        all_rows.append(row_data)
    
    # Filter out models to remove
    new_rows = []
    target_row_data = None
    for row_data in all_rows:
        model_name = str(row_data[1]).strip() if len(row_data) > 1 else ""
        if model_name in remove_names or model_name == target_model:
            if not target_row_data:
                target_row_data = row_data
                target_row_data[1] = target_model # Ensure name is correct
            continue
        new_rows.append(row_data)
    
    if not target_row_data:
        target_row_data = ['N/A', target_model] + [0] * (len(headers) - 2)
    
    # 2. Find insertion point
    insert_idx = len(new_rows) # Default to end
    for i, row_data in enumerate(new_rows):
        if str(row_data[1]).strip() == after_model:
            insert_idx = i + 1
            break
            
    new_rows.insert(insert_idx, target_row_data)
    
    # 3. Write back to sheet
    # Clear existing data (rows 2 to max)
    for row in range(ws.max_row, 1, -1):
        ws.delete_rows(row)
        
    for row_data in new_rows:
        ws.append(row_data)
    
    print(f"Updated {sheet_name}: placed {target_model} after {after_model}")

# Update 工作台: SW-4350 after SW-426
move_or_add_model('工作台', 'SW-4350', 'SW-426', ['SW4350', 'sw4350', 'SW4320'])

# Update 立柱: LG-1370-NEO+200H after LG-1000-NEO+100H
move_or_add_model('立柱', 'LG-1370-NEO+200H', 'LG-1000-NEO+100H', ['LG1370neo+200H', 'lg1370NEO+200h'])

# Update 總數: 
# SW-4350 is already there, just need to remove my duplicate SW4350
ws_summary = wb['總數']
model_col = 1
for col in range(1, ws_summary.max_column + 1):
    if ws_summary.cell(row=1, column=col).value == '機型':
        model_col = col
        break

rows_to_delete = []
for row in range(2, ws_summary.max_row + 1):
    model_name = str(ws_summary.cell(row=row, column=model_col).value).strip()
    if model_name in ['SW4350', 'SW4320', 'LG1370neo+200H']:
        rows_to_delete.append(row)

for row in reversed(rows_to_delete):
    ws_summary.delete_rows(row)

# Also check if LG-1370-NEO+200H should be added to 總數
found = False
after_row = ws_summary.max_row
for row in range(2, ws_summary.max_row + 1):
    model_name = str(ws_summary.cell(row=row, column=model_col).value).strip()
    if model_name == 'LG-1370-NEO+200H' or model_name == 'LG-1370_NEO+200H':
        found = True
        break
    if model_name == 'LG-1000_NEO+100H' or model_name == 'LG-1000-NEO+100H':
        after_row = row

if not found:
    # Use underscore style for summary if that's what others use
    target_summary_name = 'LG-1370_NEO+200H'
    ws_summary.insert_rows(after_row + 1)
    ws_summary.cell(row=after_row + 1, column=model_col, value=target_summary_name)
    print(f"Added {target_summary_name} to 總數 after LG-1000_NEO+100H")

wb.save(file_path)
print("Excel file re-ordered and names matched.")
