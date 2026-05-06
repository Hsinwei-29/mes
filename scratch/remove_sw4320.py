from openpyxl import load_workbook

file_path = '/home/hsinwei/app/mes/鑄件盤點資料.xlsx'
wb = load_workbook(file_path)

# Remove SW4320 from 工作台 sheet
ws_workbench = wb['工作台']
rows_to_delete = []
for row in range(2, ws_workbench.max_row + 1):
    if str(ws_workbench.cell(row=row, column=2).value).strip().upper() == 'SW4320':
        rows_to_delete.append(row)

for row in reversed(rows_to_delete):
    ws_workbench.delete_rows(row)
    print(f"Removed SW4320 from 工作台 row {row}")

# Remove SW4320 from 總數 sheet
ws_summary = wb['總數']
model_col = 1
for col in range(1, ws_summary.max_column + 1):
    if ws_summary.cell(row=1, column=col).value == '機型':
        model_col = col
        break

rows_to_delete = []
for row in range(2, ws_summary.max_row + 1):
    if str(ws_summary.cell(row=row, column=model_col).value).strip().upper() == 'SW4320':
        rows_to_delete.append(row)

for row in reversed(rows_to_delete):
    ws_summary.delete_rows(row)
    print(f"Removed SW4320 from 總數 row {row}")

wb.save(file_path)
print("Excel file updated.")
