from openpyxl import load_workbook

file_path = '/home/hsinwei/app/mes/鑄件盤點資料.xlsx'
wb = load_workbook(file_path)
ws = wb['總數']

seen_models = set()
rows_to_delete = []

# Assuming '機型' is in column 1 (A)
for row in range(2, ws.max_row + 1):
    model_name = ws.cell(row=row, column=1).value
    if model_name:
        model_str = str(model_name).strip()
        if model_str in seen_models:
            rows_to_delete.append(row)
            print(f"Found duplicate: {model_str} at row {row}")
        else:
            seen_models.add(model_str)

for row in reversed(rows_to_delete):
    ws.delete_rows(row)
    print(f"Deleted row {row}")

wb.save(file_path)
print("Duplicates removed from Summary sheet.")
