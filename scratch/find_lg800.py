from openpyxl import load_workbook

file_path = '/home/hsinwei/app/mes/鑄件盤點資料.xlsx'
wb = load_workbook(file_path)

print("Sheets in workbook:", wb.sheetnames)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                val_str = str(val).strip()
                if 'LG' in val_str and '800' in val_str:
                    print(f"Sheet: [{sheetname}], Row: {row}, Col: {col}, Value: {val}, Part Number (Col 1): {ws.cell(row=row, column=1).value}")
