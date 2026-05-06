from openpyxl import load_workbook

file_path = '/home/hsinwei/app/mes/鑄件盤點資料.xlsx'
wb = load_workbook(file_path)

# Update 工作台 sheet
ws_workbench = wb['工作台']
# Find if SW4350 exists
found = False
for row in range(2, ws_workbench.max_row + 1):
    if str(ws_workbench.cell(row=row, column=2).value).strip().upper() == 'SW4350':
        found = True
        break
if not found:
    new_row = ws_workbench.max_row + 1
    ws_workbench.cell(row=new_row, column=1, value='N/A')
    ws_workbench.cell(row=new_row, column=2, value='SW4350')
    # Initialize other columns with 0
    for col in range(3, ws_workbench.max_column + 1):
        ws_workbench.cell(row=new_row, column=col, value=0)
    print("Added SW4350 to 工作台")

wb.save(file_path)
print("Excel file updated.")
